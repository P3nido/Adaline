import random
import time
from pathlib import Path
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


coeficiente_aprendizado = 0.0025
epsilon = 1e-6
max_epocas = 1000
i_treinamento = 5
intervalo_entre_treinamentos_segundos = 2
NOMES_PLANILHA_TREINAMENTO = (
    "Registro de treinamento - Adalaine.xlsx",
    "Registro de treinamento - Adaline.xlsx",
)


DADOS_TREINAMENTO = [
    [0.4329, -1.3719, 0.7022, -0.8535, 1.0],
    [0.3024, 0.2286, 0.8630, 2.7909, -1.0],
    [0.1349, -0.6445, 1.0530, 0.5687, -1.0],
    [0.3374, -1.7163, 0.3670, -0.6283, -1.0],
    [1.1434, -0.0485, 0.6637, 1.2606, 1.0],
    [1.3749, -0.5071, 0.4464, 1.3009, 1.0],
    [0.7221, -0.7587, 0.7681, -0.5592, 1.0],
    [0.4403, -0.8072, 0.5154, -0.3129, 1.0],
    [-0.5231, 0.3548, 0.2538, 1.5776, -1.0],
    [0.3255, -2.0000, 0.7112, -1.1209, 1.0],
    [0.5824, 1.3915, -0.2291, 4.1735, -1.0],
    [0.1340, 0.6081, 0.4450, 3.2230, -1.0],
    [0.1480, -0.2988, 0.4778, 0.8649, 1.0],
    [0.7359, 0.1869, -0.0872, 2.3584, 1.0],
    [0.7115, -1.1469, 0.3394, 0.9573, -1.0],
    [0.8251, -1.2840, 0.8452, 1.2382, -1.0],
    [0.1569, 0.3712, 0.8825, 1.7633, 1.0],
    [0.0033, 0.6835, 0.5389, 2.8249, -1.0],
    [0.4243, 0.8313, 0.2634, 3.5855, -1.0],
    [1.0490, 0.1326, 0.9138, 1.9792, 1.0],
    [1.4276, 0.5331, -0.0145, 3.7286, 1.0],
    [0.5971, 1.4865, 0.2904, 4.6069, -1.0],
    [0.8475, 2.1479, 0.3179, 5.8235, -1.0],
    [1.3967, -0.4171, 0.6443, 1.3927, 1.0],
    [0.0044, 1.5378, 0.6099, 4.7755, -1.0],
    [0.2201, -0.5668, 0.0515, 0.7829, 1.0],
    [0.6300, -1.2480, 0.8591, 0.8093, -1.0],
    [-0.2479, 0.8960, 0.0547, 1.7381, 1.0],
    [-0.3088, -0.0929, 0.8659, 1.5483, -1.0],
    [-0.5180, 1.4974, 0.5453, 2.3993, -1.0],
    [0.6833, 0.8266, 0.0829, 2.8864, 1.0],
    [0.4353, -1.4066, 0.4207, -0.4879, 1.0],
    [-0.1069, -3.2329, 0.1856, -2.4572, -1.0],
    [0.4662, 0.6261, 0.7304, 3.4370, -1.0],
    [0.8298, -1.4089, 0.3119, 1.3235, -1.0],
]


DADOS_VALIDACAO = [
    [0.9694, 0.6909, 0.4334, 3.4965],
    [0.5427, 1.3832, 0.6390, 4.0352],
    [0.6081, -0.9196, 0.5925, 0.1016],
    [-0.1618, 0.4694, 0.2030, 3.0117],
    [0.1870, -0.2578, 0.6124, 1.7749],
    [0.4891, -0.5276, 0.4378, 0.6439],
    [0.3777, 2.0149, 0.7423, 3.3932],
    [1.1498, -0.4067, 0.2469, 1.5866],
    [0.9325, 1.0950, 1.0359, 3.3591],
    [0.5060, 1.3317, 0.9222, 3.7174],
    [0.0497, -2.0656, 0.6124, -0.6585],
    [0.4004, 3.5369, 0.9766, 5.3532],
    [-0.1874, 1.3343, 0.5374, 3.2189],
    [0.5060, 1.3317, 0.9222, 3.7174],
    [1.6375, -0.7911, 0.7537, 0.5515],
]


def carregar_dados_de_vetores():
    """Separa dados de treino/validacao dos vetores e adiciona bias = -1."""
    dados_treinamento = np.array(DADOS_TREINAMENTO, dtype=float)
    dados_validacao = np.array(DADOS_VALIDACAO, dtype=float)

    x_treino_sem_bias = dados_treinamento[:, :4]
    d_treino = dados_treinamento[:, 4]

    bias_treino = -np.ones((x_treino_sem_bias.shape[0], 1), dtype=float)
    x_treino = np.hstack((bias_treino, x_treino_sem_bias))

    bias_validacao = -np.ones((dados_validacao.shape[0], 1), dtype=float)
    x_validacao = np.hstack((bias_validacao, dados_validacao))

    return x_treino, d_treino, x_validacao


def saida_linear(pesos, amostra):
    """Calcula a saida linear u = sum(w_i * x_i)."""
    return float(np.dot(pesos, amostra))


def calcular_eqm_online(soma_erros_quadrados, total_amostras):
    """Calcula o Erro Quadratico Medio da epoca."""
    return soma_erros_quadrados / total_amostras


def treinar_adaline_online(x, d, taxa_aprendizado, precisao_eqm, max_epocas, pesos_iniciais):
    """Treina Adaline no modo online (amostra por amostra)."""
    pesos = pesos_iniciais.copy().astype(float)
    eqm_anterior = np.inf
    historico_eqm = []
    epoca = 0

    while epoca < max_epocas:
        soma_erros_quadrados = 0.0

        for amostra, desejado in zip(x, d):
            u = saida_linear(pesos, amostra)
            erro = desejado - u

            # Regra Delta (atualizacao online)
            pesos = pesos + taxa_aprendizado * erro * amostra
            soma_erros_quadrados += erro ** 2

        eqm_atual = calcular_eqm_online(soma_erros_quadrados, x.shape[0])
        historico_eqm.append(eqm_atual)
        epoca += 1

        if abs(eqm_atual - eqm_anterior) <= precisao_eqm:
            break

        eqm_anterior = eqm_atual

    return pesos, epoca, historico_eqm


def gerar_pesos_iniciais(tamanho_vetor, numero_treinamento):
    """Gera pesos iniciais em [0, 1] com seed baseada no timestamp atual."""
    seed_atual = time.time_ns() + numero_treinamento
    random.seed(seed_atual)
    pesos = np.array([random.random() for _ in range(tamanho_vetor)], dtype=float)
    return pesos, seed_atual


def formatar_vetor(vetor):
    return "[" + ", ".join(f"{valor:.6f}" for valor in vetor) + "]"


def imprimir_tabela_resultados(resultados):
    cabecalho = [
        "Treinamento",
        "Pesos iniciais (w0...w4)",
        "Pesos finais (w0...w4)",
        "Epocas",
    ]

    linhas = []
    for resultado in resultados:
        linhas.append(
            [
                str(resultado["treinamento"]),
                formatar_vetor(resultado["pesos_iniciais"]),
                formatar_vetor(resultado["pesos_finais"]),
                str(resultado["epocas"]),
            ]
        )

    larguras = [len(coluna) for coluna in cabecalho]
    for linha in linhas:
        for i, valor in enumerate(linha):
            larguras[i] = max(larguras[i], len(valor))

    separador = "-+-".join("-" * largura for largura in larguras)
    linha_cabecalho = " | ".join(
        titulo.ljust(larguras[i]) for i, titulo in enumerate(cabecalho)
    )

    print("\nRESULTADOS DOS 5 TREINAMENTOS (ADALINE)")
    print(linha_cabecalho)
    print(separador)

    for linha in linhas:
        print(" | ".join(linha[i].ljust(larguras[i]) for i in range(len(cabecalho))))


def plotar_eqm(historico_eqm):
    """Funcao pronta para uso futuro com matplotlib."""
    plt.figure(figsize=(8, 4))
    plt.plot(range(1, len(historico_eqm) + 1), historico_eqm)
    plt.title("Evolucao do EQM")
    plt.xlabel("Epoca")
    plt.ylabel("EQM")
    plt.grid(True)
    plt.tight_layout()
    plt.close()


def preencher_pesos_na_linha(folha, linha, coluna_inicial, coluna_final, pesos):
    for deslocamento, coluna in enumerate(range(coluna_inicial, coluna_final + 1)):
        valor = float(pesos[deslocamento]) if deslocamento < len(pesos) else None
        celula = folha.cell(row=linha, column=coluna)
        celula.value = valor
        if valor is not None:
            celula.number_format = "0.000000"


def registrar_treinamentos_em_planilha(resultados, caminho_planilha):
    if not caminho_planilha.exists():
        raise FileNotFoundError(f"Planilha de registro nao encontrada: {caminho_planilha}")

    workbook = load_workbook(caminho_planilha)
    folha = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active

    linha_inicial = 8
    linha_final = linha_inicial + i_treinamento - 1

    # Limpa os valores antigos da area de resultados e preserva o layout do template.
    for linha in range(linha_inicial, linha_final + 1):
        for coluna in range(6, 19):
            folha.cell(row=linha, column=coluna).value = None

    for indice, resultado in enumerate(resultados, start=1):
        linha = linha_inicial + indice - 1

        folha.cell(row=linha, column=5).value = f"#{indice} T({indice})"
        preencher_pesos_na_linha(
            folha,
            linha,
            coluna_inicial=6,
            coluna_final=11,
            pesos=resultado["pesos_iniciais"],
        )
        preencher_pesos_na_linha(
            folha,
            linha,
            coluna_inicial=12,
            coluna_final=17,
            pesos=resultado["pesos_finais"],
        )
        folha.cell(row=linha, column=18).value = int(resultado["epocas"])

    workbook.save(caminho_planilha)


def resolver_caminho_planilha_treinamento(pasta_projeto):
    for nome_arquivo in NOMES_PLANILHA_TREINAMENTO:
        caminho = pasta_projeto / nome_arquivo
        if caminho.exists():
            return caminho

    return pasta_projeto / NOMES_PLANILHA_TREINAMENTO[-1]


def main():
    pasta_projeto = Path(__file__).resolve().parent
    x, d, x_validacao = carregar_dados_de_vetores()
    caminho_planilha_treinamento = resolver_caminho_planilha_treinamento(pasta_projeto)

    # Mantido para uso futuro do projeto (relatorios).
    _ = x_validacao

    resultados = []
    for treino in range(1, i_treinamento + 1):
        pesos_iniciais, seed_usada = gerar_pesos_iniciais(x.shape[1], treino)

        pesos_finais, epocas, historico_eqm = treinar_adaline_online(
            x=x,
            d=d,
            taxa_aprendizado=coeficiente_aprendizado,
            precisao_eqm=epsilon,
            max_epocas=max_epocas,
            pesos_iniciais=pesos_iniciais,
        )

        resultados.append(
            {
                "treinamento": treino,
                "seed": seed_usada,
                "pesos_iniciais": pesos_iniciais,
                "pesos_finais": pesos_finais,
                "epocas": epocas,
                "eqm_final": historico_eqm[-1],
            }
        )
        print(f"Treinamento {treino} concluido")

        if treino < i_treinamento:
            time.sleep(intervalo_entre_treinamentos_segundos)

    imprimir_tabela_resultados(resultados)
    registrar_treinamentos_em_planilha(resultados, caminho_planilha_treinamento)


if __name__ == "__main__":
    main()